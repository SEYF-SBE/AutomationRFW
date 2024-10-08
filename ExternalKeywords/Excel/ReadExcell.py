import openpyxl

# open Workbook

wk = openpyxl.load_workbook('ExternalKeywords/Excel/TestData.xlsx')
print(wk.sheetnames)
print("Active sheet is : " + wk.active.title)

# Create object of any sheet  on which i want to work

sht = wk['LoginData']
print(sht.title)

# read a cell
# print(sht['A5'].value)
#c1 = sht.cell(2, 2)
#print(c1.value)

def fetch_number_of_rows(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname, row, cell):
    sh = wk[Sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value

# find rows having data
# rows = sht.max_row
# cols = sht.max_column
# print(f"rows : {rows}, and cols : {cols}" )
# for i in range(1, rows+1):
    # for j in range(1,cols+1):
      # print(sht.cell(i,j).value)

# 2nd approche without know how many rows and cols we have
# for r in sht['A1':'C4']:
   # for c in r:
        # print(c.value)


